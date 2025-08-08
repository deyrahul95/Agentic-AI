from typing import Dict, List, Optional
import openpyxl
import uuid

from google.adk.tools import ToolContext
from openpyxl.worksheet.worksheet import Worksheet

from task_tracker_agent.models import (
    PriorityLevel,
    TaskStatus,
    Task,
    AddTaskResponse,
    UpdateDeleteResponse,
    ReadTaskResponse,
)


EXCEL_PATH = "data.xlsx"
SHEET_NAME = "Tasks"
PRIORITY_ORDER: Dict[PriorityLevel, int] = {"High": 1, "Medium": 2, "Low": 3}


def init_sheet() -> None:
    """Initialize empty excel sheet"""
    try:
        wb = openpyxl.load_workbook(EXCEL_PATH)
    except FileNotFoundError:
        wb = openpyxl.Workbook()
    if SHEET_NAME not in wb.sheetnames:
        ws = wb.active
        if ws is None:
            raise Exception("Active workbook is missing!")
        ws.title = SHEET_NAME
        ws.append(["ID", "Task", "Priority", "Status"])
    wb.save(EXCEL_PATH)


def sort_tasks(ws: Worksheet):
    """Sort tasks in-place by priority order."""
    data = list(ws.iter_rows(min_row=2, values_only=True))
    data.sort(key=lambda row: PRIORITY_ORDER.get(row[2], 3))  # type: ignore
    ws.delete_rows(2, ws.max_row)
    for row in data:
        ws.append(row)


def parse_priority(description: str) -> PriorityLevel:
    """
    Infer priority level from task description.

    Args:
        description (str): User-provided task description.
    Returns:
        str: One of "High", "Medium", or "Low".
    """
    desc = description.lower()
    if any(
        term in desc
        for term in ["urgent", "immediately", "asap", "deadline", "important"]
    ):
        return "High"
    if any(term in desc for term in ["later", "whenever", "eventually", "sometime"]):
        return "Low"
    return "Medium"


def add_task(tool_context: ToolContext, description: str) -> AddTaskResponse:
    """
    Adds a new task, inferring its priority automatically.

    Args:
        tool_context (ToolContext): ADK tool context.
        description (str): The description of the task to add.
    Returns:
        Dict: Contains 'status' and 'task' summary.
    """
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb[SHEET_NAME]
    task_id = str(uuid.uuid4())[:8]
    priority = parse_priority(description)
    ws.append([task_id, description, priority, "Incomplete"])
    sort_tasks(ws)
    wb.save(EXCEL_PATH)
    return {
        "status": "success",
        "task": Task(
            id=task_id, description=description, priority=priority, status="Incomplete"
        ),
    }


def read_tasks(tool_context: ToolContext, query: str = "") -> ReadTaskResponse:
    """
    Reads tasks filtered by an optional query keyword.

    Args:
        tool_context (ToolContext): ADK tool context.
        query (str, optional): Keyword to filter tasks.
    Returns:
        Dict: Contains 'tasks' (list of matching task dicts).
    """
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb[SHEET_NAME]
    headers = [cell.value for cell in ws[1]]
    tasks: List[Task] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        row_dict = dict(zip(headers, row))
        if not query or query.lower() in str(row_dict["Task"]).lower():
            task = Task(
                id=str(row_dict["ID"]),
                description=str(row_dict["Task"]),
                priority=row_dict["Priority"],  # type: ignore
                status=row_dict["Status"],  # type: ignore
            )
            tasks.append(task)

    return {"status": "success", "tasks": tasks}


def update_task(
    tool_context: ToolContext,
    query: str,
    new_description: Optional[str] = None,
    new_status: Optional[TaskStatus] = None,
) -> UpdateDeleteResponse:
    """
    Updates a task's description or status.

    Args:
        tool_context (ToolContext): ADK tool context.
        query (str): Keyword to identify the task.
        new_description (str, optional): Updated description.
        new_status (str, optional): Updated status ("Complete" or "Incomplete").
    Returns:
        Dict: Contains 'status' and an 'updated' flag.
    """
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb[SHEET_NAME]
    updated = False

    for row in ws.iter_rows(min_row=2):
        if query.lower() in str(row[1].value).lower():
            if new_description:
                row[1].value = new_description  # type: ignore
            if new_status:
                row[3].value = new_status.capitalize()  # type: ignore
            updated = True

    if updated:
        sort_tasks(ws)
        wb.save(EXCEL_PATH)
        return {"status": "success", "updated": True, "deleted": None, "message": None}
    return {
        "status": "error",
        "updated": False,
        "deleted": None,
        "message": "Task not found",
    }


def delete_task(tool_context: ToolContext, query: str) -> UpdateDeleteResponse:
    """
    Deletes a task by keyword.

    Args:
        tool_context (ToolContext): ADK tool context.
        query (str): Keyword to identify the task for deletion.
    Returns:
        Dict: Contains 'status' and 'deleted' flag.
    """
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb[SHEET_NAME]
    deleted = False

    for i in range(ws.max_row, 1, -1):
        task_cell = ws.cell(row=i, column=2)
        if query.lower() in str(task_cell.value).lower():
            ws.delete_rows(i)
            deleted = True

    if deleted:
        wb.save(EXCEL_PATH)
        return {"status": "success", "deleted": True, "updated": None, "message": None}
    return {
        "status": "error",
        "deleted": False,
        "updated": None,
        "message": "Task not found",
    }
